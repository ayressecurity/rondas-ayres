"""
Informes del libro de novedades (solo lectura). Cada vista reutiliza la base
común (apps/informes/base.py) y solo define su filtro de tipo_evento. La
exportación a Excel respeta el filtro de instalación Y el filtro de fechas.
"""
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.comun.decoradores import requiere_instalacion
from apps.novedades.models import CategoriaEvento, LibroNovedadesMedia, TipoMedia
from .base import eventos_filtrados, render_informe


def _filtro_rondas(qs):
    """Todos los eventos MENOS los de categoría 'novedad'."""
    return qs.exclude(tipo_evento__categoria=CategoriaEvento.NOVEDAD)


def _filtro_novedades(qs):
    """ÚNICAMENTE eventos de categoría 'novedad'."""
    return qs.filter(tipo_evento__categoria=CategoriaEvento.NOVEDAD)


@login_required
@requiere_instalacion
def informe_rondas(request):
    return render_informe(
        request,
        titulo="Informe de Rondas",
        aplica_filtro=_filtro_rondas,
        export_url="informes:exportar_rondas",
    )


@login_required
@requiere_instalacion
def informe_novedades(request):
    return render_informe(
        request,
        titulo="Informe de Novedades",
        aplica_filtro=_filtro_novedades,
        export_url="informes:exportar_novedades",
        template="informes/novedades.html",
        con_imagen=True,
    )


# Color de marca (sin '#') para openpyxl.
ROJO_MARCA = "CC3333"
COLUMNAS = [
    ("Fecha/Hora", 20),
    ("Tipo de evento", 22),
    ("Punto de control", 22),
    ("Guardia", 24),
    ("Coordenadas", 30),
    ("Distancia (m)", 14),
    ("Geocerca", 12),
    ("Observación", 40),
]


def _geocerca_texto(ev):
    if ev.dentro_geocerca is True:
        return "Dentro"
    if ev.dentro_geocerca is False:
        return "Fuera"
    return "—"


# Color de hipervínculo de Excel (azul estándar).
AZUL_LINK = "0563C1"


def _exportar_excel(request, *, titulo, aplica_filtro, slug_base):
    """Genera el .xlsx del informe respetando instalación + filtro de fechas.
    La columna Coordenadas va como HIPERVÍNCULO a Google Maps (clickeable)."""
    eventos, _rango, etiqueta, _valores = eventos_filtrados(request, aplica_filtro)
    instalacion = request.session.get("instalacion_nombre") or "instalacion"

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita el nombre de hoja a 31 chars

    # Título (fila 1, combinada) con instalación y rango de fechas filtrado.
    n_cols = len(COLUMNAS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    celda_titulo = ws.cell(row=1, column=1, value=f"{titulo} — {instalacion} — {etiqueta}")
    celda_titulo.font = Font(bold=True, size=13, color=ROJO_MARCA)
    celda_titulo.alignment = Alignment(horizontal="left", vertical="center")

    # Encabezados (fila 2): negrita, fondo rojo de marca, texto blanco.
    relleno = PatternFill(start_color=ROJO_MARCA, end_color=ROJO_MARCA, fill_type="solid")
    fuente = Font(bold=True, color="FFFFFF")
    centrado = Alignment(horizontal="center", vertical="center")
    fila_encabezado = 2
    for col, (nombre, ancho) in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=nombre)
        celda.fill = relleno
        celda.font = fuente
        celda.alignment = centrado
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Datos.
    fila = fila_encabezado + 1
    for ev in eventos:
        ws.cell(row=fila, column=1,
                value=timezone.localtime(ev.timestamp_evento).strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=fila, column=2, value=ev.tipo_evento.nombre)
        ws.cell(row=fila, column=3, value=ev.punto_control.nombre if ev.punto_control else "—")
        ws.cell(row=fila, column=4, value=ev.guardia_nombre)

        # Coordenadas como hipervínculo a Google Maps (igual que "Ver mapa").
        celda_coord = ws.cell(row=fila, column=5)
        if ev.lat is not None and ev.lng is not None:
            url = f"https://www.google.com/maps?q={ev.lat},{ev.lng}"  # punto decimal, completo
            celda_coord.value = "Ver mapa"
            celda_coord.hyperlink = url
            celda_coord.font = Font(color=AZUL_LINK, underline="single")
        else:
            celda_coord.value = "—"

        ws.cell(row=fila, column=6,
                value=float(ev.distancia_metros) if ev.distancia_metros is not None else "—")
        ws.cell(row=fila, column=7, value=_geocerca_texto(ev))
        ws.cell(row=fila, column=8, value=ev.texto or "—")
        fila += 1

    ws.freeze_panes = "A3"  # fija título + encabezado al hacer scroll

    hoy = timezone.localtime(timezone.now()).date().isoformat()
    nombre_archivo = f"{slug_base}_{slugify(instalacion)}_{hoy}.xlsx"

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(resp)
    return resp


@login_required
@requiere_instalacion
def exportar_rondas(request):
    """Exporta el Informe de Rondas a .xlsx respetando instalación + fechas."""
    return _exportar_excel(
        request, titulo="Informe de Rondas", aplica_filtro=_filtro_rondas, slug_base="informe_rondas"
    )


# --- Excel de Novedades: columnas propias + imágenes insertadas ---
COLUMNAS_NOVEDAD = [
    ("Fecha/Hora", 20),
    ("Observación", 45),
    ("Tipo de evento", 20),
    ("Guardia", 24),
]
IMG_PX = 90          # tamaño uniforme de cada miniatura en el Excel
COL_IMG_ANCHO = 14   # ancho de columna que encaja ~90px
FILA_ALTO = 70       # alto de fila que encaja ~90px


def _fotos_por_evento(eventos):
    """{libro_id: [paths...]} de las fotos de los eventos dados (1 query)."""
    ids = [ev.id for ev in eventos]
    fotos = {}
    if ids:
        for libro_id, path in (
            LibroNovedadesMedia.objects
            .filter(libro_novedades_id__in=ids, tipo=TipoMedia.FOTO)
            .order_by("id")
            .values_list("libro_novedades_id", "path")
        ):
            fotos.setdefault(libro_id, []).append(path)
    return fotos


@login_required
@requiere_instalacion
def exportar_novedades(request):
    """Exporta el Informe de Novedades a .xlsx (respeta instalación + fechas),
    INSERTANDO las fotos reales en la columna de imágenes."""
    eventos, _rango, etiqueta, _valores = eventos_filtrados(request, _filtro_novedades)
    instalacion = request.session.get("instalacion_nombre") or "instalacion"
    fotos = _fotos_por_evento(eventos)
    max_fotos = max((len(v) for v in fotos.values()), default=0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Informe de Novedades"[:31]

    base_cols = len(COLUMNAS_NOVEDAD)
    total_cols = base_cols + max(max_fotos, 1)  # +columnas de imágenes

    # Título (fila 1) con instalación + rango.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    celda_titulo = ws.cell(row=1, column=1, value=f"Informe de Novedades — {instalacion} — {etiqueta}")
    celda_titulo.font = Font(bold=True, size=13, color=ROJO_MARCA)
    celda_titulo.alignment = Alignment(horizontal="left", vertical="center")

    # Encabezados (fila 2): rojo de marca + blanco.
    relleno = PatternFill(start_color=ROJO_MARCA, end_color=ROJO_MARCA, fill_type="solid")
    fuente = Font(bold=True, color="FFFFFF")
    centrado = Alignment(horizontal="center", vertical="center")
    for col, (nombre, ancho) in enumerate(COLUMNAS_NOVEDAD, start=1):
        celda = ws.cell(row=2, column=col, value=nombre)
        celda.fill = relleno
        celda.font = fuente
        celda.alignment = centrado
        ws.column_dimensions[get_column_letter(col)].width = ancho
    # Encabezado de imágenes: la banda roja cubre TODAS las columnas de imágenes
    # (la novedad con más fotos define el ancho), no solo la primera.
    for c in range(base_cols + 1, total_cols + 1):
        celda = ws.cell(row=2, column=c)
        celda.fill = relleno
        celda.font = fuente
        celda.alignment = centrado
        ws.column_dimensions[get_column_letter(c)].width = COL_IMG_ANCHO
    ws.cell(row=2, column=base_cols + 1, value="Imágenes")
    if total_cols > base_cols + 1:
        ws.merge_cells(start_row=2, start_column=base_cols + 1, end_row=2, end_column=total_cols)

    # Datos.
    fila = 3
    for ev in eventos:
        ws.cell(row=fila, column=1,
                value=timezone.localtime(ev.timestamp_evento).strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=fila, column=2, value=ev.texto or "—")
        ws.cell(row=fila, column=3, value=ev.tipo_evento.nombre)
        ws.cell(row=fila, column=4, value=ev.guardia_nombre)

        paths = fotos.get(ev.id, [])
        if paths:
            ws.row_dimensions[fila].height = FILA_ALTO
            for j, path in enumerate(paths):
                try:
                    with default_storage.open(path) as fh:
                        datos = fh.read()
                except (FileNotFoundError, OSError):
                    continue
                imagen = XLImage(BytesIO(datos))
                imagen.width = IMG_PX
                imagen.height = IMG_PX
                # Cada foto en una columna contigua a partir de la de imágenes.
                ancla = ws.cell(row=fila, column=base_cols + 1 + j).coordinate
                ws.add_image(imagen, ancla)
        else:
            ws.cell(row=fila, column=base_cols + 1, value="—")
        fila += 1

    ws.freeze_panes = "A3"

    hoy = timezone.localtime(timezone.now()).date().isoformat()
    nombre_archivo = f"informe_novedades_{slugify(instalacion)}_{hoy}.xlsx"

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(resp)
    return resp
