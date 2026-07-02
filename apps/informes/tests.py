"""
Tests de la resolución de filtros de fecha de los informes (apps/informes/base.py).

Solo dos filtros: AÑO y RANGO (fini/ffin). El día actual es un DEFAULT AUTOMÁTICO
e INVISIBLE: se aplica solo cuando no hay año ni rango. Si el usuario usa año o
rango, ese filtro manda y "hoy" NO interviene.
"""
import jwt
from django.contrib.auth import get_user_model
from django.test import Client, RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from apps.espejo.models import Cliente, Instalacion
from apps.informes.base import _rango_y_label
from apps.novedades.models import CategoriaEvento, LibroNovedades, TipoEvento


class RangoFechaDefaultTests(TestCase):
    def setUp(self):
        self.rf = RequestFactory()

    def test_sin_parametros_default_hoy(self):
        request = self.rf.get("/informes/rondas/")  # primera carga, sin querystring
        rango, etiqueta, valores = _rango_y_label(request)

        hoy = timezone.localtime(timezone.now()).date()
        self.assertIsNotNone(rango)
        self.assertEqual(etiqueta, f"Día {hoy.isoformat()}")
        # 'hoy' es default automático: NO hay control de día (no se repinta).
        self.assertNotIn("dia", valores)

    def test_formulario_vacio_tambien_default_hoy(self):
        # Año y rango vacíos = no hay filtro activo -> hoy.
        request = self.rf.get("/informes/rondas/", {"anio": "", "fini": "", "ffin": ""})
        rango, etiqueta, _ = _rango_y_label(request)

        hoy = timezone.localtime(timezone.now()).date()
        self.assertIsNotNone(rango)
        self.assertEqual(etiqueta, f"Día {hoy.isoformat()}")

    def test_rango_se_respeta_sin_forzar_hoy(self):
        # Rango que abarca de un mes a otro (calendario).
        request = self.rf.get(
            "/informes/rondas/", {"fini": "2026-03-01", "ffin": "2026-05-31"}
        )
        rango, etiqueta, valores = _rango_y_label(request)

        self.assertEqual(etiqueta, "2026-03-01 a 2026-05-31")
        inicio, fin = rango
        # [2026-03-01, 2026-06-01) -> 92 días (marzo 31 + abril 30 + mayo 31).
        self.assertEqual((fin - inicio).days, 92)
        self.assertEqual(valores["fini"], "2026-03-01")
        self.assertEqual(valores["ffin"], "2026-05-31")

    def test_anio_se_respeta_sin_forzar_hoy(self):
        request = self.rf.get("/informes/rondas/", {"anio": "2026"})
        rango, etiqueta, valores = _rango_y_label(request)

        self.assertIsNotNone(rango)
        self.assertEqual(etiqueta, "Año 2026")
        self.assertEqual(valores["anio"], "2026")

    def test_anio_tiene_precedencia_sobre_rango(self):
        # Si por algún motivo vienen ambos, manda el año (precedencia definida).
        request = self.rf.get(
            "/informes/rondas/",
            {"anio": "2026", "fini": "2026-03-01", "ffin": "2026-05-31"},
        )
        _rango, etiqueta, _valores = _rango_y_label(request)
        self.assertEqual(etiqueta, "Año 2026")


class ExportExcelSaneoTests(TestCase):
    """El export no debe romper con caracteres de control en el texto (QA #1)."""

    def setUp(self):
        self.user = get_user_model().objects.create(username="u")
        self.client = Client()
        self.client.force_login(self.user)
        s = self.client.session
        s["cliente_id"] = 1
        s["instalacion_id"] = 10
        s["instalacion_nombre"] = "Inst"
        s.save()
        tipo = TipoEvento.objects.create(codigo="arribo", nombre="Arribo", categoria=CategoriaEvento.RONDA)
        ahora = timezone.now()  # cae en el default "hoy" del informe
        LibroNovedades.objects.create(
            instalacion_id=10, guardia_keycloak_id="x", tipo_evento=tipo,
            timestamp_evento=ahora, timestamp_servidor=ahora, estado="ok",
            texto="texto\x07con\x00control",  # chars ilegales para openpyxl
        )

    def test_export_rondas_no_rompe_con_char_de_control(self):
        resp = self.client.get("/informes/rondas/excel/")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])


@override_settings(STORAGES={
    "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
    "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
})
class ComentarioCentralInformesTests(TestCase):
    """Columna 'Comentario' (comentario_central) en ambos informes: pantalla y Excel.

    El campo es de solo-lectura para los informes: se muestra tal cual está en el
    evento (lo escribe la Central de Monitoreo). NULL -> '—' en pantalla / '' en Excel.
    """

    def setUp(self):
        self.user = get_user_model().objects.create(username="u")
        self.client = Client()
        self.client.force_login(self.user)
        s = self.client.session
        s["cliente_id"] = 1
        s["instalacion_id"] = 10
        s["instalacion_nombre"] = "Inst"
        s.save()

    def _evento(self, tipo, **extra):
        # timestamp_evento = ahora -> cae en el default "hoy" del informe.
        ahora = timezone.now()
        return LibroNovedades.objects.create(
            instalacion_id=10, guardia_keycloak_id="x", tipo_evento=tipo,
            timestamp_evento=ahora, timestamp_servidor=ahora, estado="ok", **extra,
        )

    # ---- pantalla ----
    def test_comentario_en_informe_rondas(self):
        # 'arribo' es categoría RONDA -> aparece en el Informe de Rondas.
        tipo = TipoEvento.objects.create(codigo="arribo", nombre="Arribo", categoria=CategoriaEvento.RONDA)
        self._evento(tipo, texto="obs", comentario_central="Revisar cámara 3")
        resp = self.client.get("/informes/rondas/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "<th>Comentario</th>")
        self.assertContains(resp, "Revisar cámara 3")

    def test_comentario_en_informe_novedades(self):
        # 'novedad' es categoría NOVEDAD -> aparece en el Informe de Novedades.
        tipo = TipoEvento.objects.create(codigo="novedad", nombre="Novedad", categoria=CategoriaEvento.NOVEDAD)
        self._evento(tipo, texto="obs", comentario_central="Comentario central X")
        resp = self.client.get("/informes/novedades/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "<th>Comentario</th>")
        self.assertContains(resp, "Comentario central X")

    def test_sin_comentario_muestra_guion(self):
        # comentario_central NULL -> la celda existe con '—' (no rompe la fila).
        tipo = TipoEvento.objects.create(codigo="arribo", nombre="Arribo", categoria=CategoriaEvento.RONDA)
        self._evento(tipo, texto="obs")
        resp = self.client.get("/informes/rondas/")
        self.assertContains(resp, 'data-col="Comentario">—')

    # ---- Excel ----
    def _valores_excel(self, resp):
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content))
        return {c.value for row in wb.active.iter_rows() for c in row}

    def test_comentario_en_excel_rondas(self):
        tipo = TipoEvento.objects.create(codigo="arribo", nombre="Arribo", categoria=CategoriaEvento.RONDA)
        self._evento(tipo, texto="obs", comentario_central="ComentarioExcelR")
        resp = self.client.get("/informes/rondas/excel/")
        self.assertEqual(resp.status_code, 200)
        valores = self._valores_excel(resp)
        self.assertIn("Comentario", valores)        # encabezado nuevo
        self.assertIn("ComentarioExcelR", valores)  # valor de la fila

    def test_comentario_en_excel_novedades(self):
        tipo = TipoEvento.objects.create(codigo="novedad", nombre="Novedad", categoria=CategoriaEvento.NOVEDAD)
        self._evento(tipo, texto="obs", comentario_central="ComentarioExcelN")
        resp = self.client.get("/informes/novedades/excel/")
        self.assertEqual(resp.status_code, 200)
        valores = self._valores_excel(resp)
        self.assertIn("Comentario", valores)
        self.assertIn("ComentarioExcelN", valores)


@override_settings(STORAGES={
    "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
    "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
})
class InformesAislamientoClienteTests(TestCase):
    """3.3a (defensa en profundidad): un rol 'cliente' no ve informes de una
    instalación de otra empresa, aunque su sesión traiga ese instalacion_id."""

    def setUp(self):
        self.user = get_user_model().objects.create(username="cli")
        self.client = Client()
        self.client.force_login(self.user)
        Cliente.objects.create(id=400, razon_social="Muni Las Condes", rut="400-9")
        Cliente.objects.create(id=500, razon_social="Otro Cliente", rut="500-9")
        Instalacion.objects.create(id=40, codigo="AYR-0040", cliente_id=400, nombre="Propia")
        Instalacion.objects.create(id=50, codigo="AYR-0050", cliente_id=500, nombre="Ajena")
        token = jwt.encode(
            {"realm_access": {"roles": ["cliente"]}, "cliente_id": "400"},
            "x", algorithm="HS256",
        )
        s = self.client.session
        s["oidc_access_token"] = token
        s["cliente_id"] = 400
        s.save()

    def _fijar_instalacion(self, ins_id, nombre):
        s = self.client.session
        s["instalacion_id"] = ins_id
        s["instalacion_nombre"] = nombre
        s.save()

    def test_instalacion_ajena_en_sesion_se_rechaza(self):
        # Sesión manipulada: instalacion de cliente 500 mientras el token es 400.
        self._fijar_instalacion(50, "Ajena")
        resp = self.client.get(reverse("informes:rondas"))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse("instalaciones:index"))
        self.assertNotIn("instalacion_id", self.client.session)   # descartada

    def test_instalacion_propia_renderiza(self):
        self._fijar_instalacion(40, "Propia")
        resp = self.client.get(reverse("informes:rondas"))
        self.assertEqual(resp.status_code, 200)


class InsercionNoTocaComentarioTests(TestCase):
    """Los flujos de inserción NO setean comentario_central: queda NULL (sin tocarse)."""

    def test_registrar_evento_simple_deja_comentario_null(self):
        from apps.comun.services.rondas import registrar_evento_simple
        TipoEvento.objects.create(codigo="novedad", nombre="Novedad", categoria=CategoriaEvento.NOVEDAD)
        res = registrar_evento_simple(
            instalacion_id=10, guardia_keycloak_id="x", codigo_tipo="novedad",
            texto="obs", ahora=timezone.now(),
        )
        ev = LibroNovedades.objects.get(id=res["libro_id"])
        self.assertIsNone(ev.comentario_central)

    def test_registrar_escaneo_deja_comentario_null(self):
        from datetime import date, time as dtime

        from apps.checkpoints.models import PuntoControl
        from apps.comun.services.rondas import registrar_escaneo
        from apps.rondas.models import Ronda

        TipoEvento.objects.create(codigo="arribo", nombre="Arribo", categoria=CategoriaEvento.RONDA)
        cp = PuntoControl.objects.create(
            instalacion_id=10, nombre="P", lat="-33.4", lng="-70.56",
            tolerancia_mts=100, validar_posicion=True,
            qr_token="cccccccc-cccc-cccc-cccc-cccccccccccc", activo=True,
        )
        # Ronda activa todo el día (contiene "ahora") y sin programación -> ventana
        # = turno completo; el escaneo dentro de tolerancia registra un 'arribo'.
        Ronda.objects.create(
            cliente_id=1, instalacion_id=10, nombre="R", fecha_inicio=date(2026, 1, 1),
            hora_inicio=dtime(0, 0), hora_fin=dtime(23, 59, 59),
        )
        res = registrar_escaneo(
            instalacion_id=10, guardia_keycloak_id="x", qr_token=cp.qr_token,
            lat=-33.4, lng=-70.56, texto=None, ahora=timezone.now(),
        )
        self.assertEqual(res["resultado"], "ok")
        ev = LibroNovedades.objects.get(id=res["libro_id"])
        self.assertIsNone(ev.comentario_central)
